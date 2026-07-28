from io import BytesIO
from datetime import datetime

from aiogram import Router, F, Bot
from aiogram.types import (
    Message, CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton, BufferedInputFile,
)
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup

from config import CATEGORIES
from db.repos import Repos
from handlers.staff.auth import require_role
from utils import safe_edit

router = Router()


def kb_manager() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📊 Общая статистика", callback_data="mgr:stats")],
        [InlineKeyboardButton(text="💰 Финансы", callback_data="mgr:finance")],
        [InlineKeyboardButton(text="📥 Экспорт в Excel", callback_data="mgr:export")],
        [InlineKeyboardButton(text="✏️ Тексты и баннер", callback_data="mgr:texts")],
        [InlineKeyboardButton(text="👁 Просмотр каналов и разделов", callback_data="mgr:overview")],
        [InlineKeyboardButton(text="📂 Разделы (топики)", callback_data="mgr:brands")],
        [InlineKeyboardButton(text="📡 Каналы (муж/жен)", callback_data="mgr:channels")],
        [InlineKeyboardButton(text="📦 Канал отчётов склада", callback_data="mgr:report_channel")],
        [InlineKeyboardButton(text="➕ Создать аккаунт", callback_data="mgr:create_account")],
        [InlineKeyboardButton(text="🗑 Удалить аккаунт", callback_data="mgr:delete_account")],
        [InlineKeyboardButton(text="👥 Список аккаунтов", callback_data="mgr:list_accounts")],
        [InlineKeyboardButton(text="🔑 Сменить свой пароль", callback_data="mgr:change_my_password")],
        [InlineKeyboardButton(text="🚪 Выйти", callback_data="mgr:logout")],
    ])


def kb_back_manager() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="⬅️ Назад", callback_data="mgr:back")]])


async def show_manager_panel(message: Message):
    await message.answer("🧑‍💼 <b>Панель менеджера</b>", reply_markup=kb_manager())


class ManagerStates(StatesGroup):
    create_role = State()
    create_login = State()
    create_password = State()
    create_commission = State()
    change_my_password = State()
    delete_login = State()
    editing_text = State()
    channel_gender = State()
    channel_chat_id = State()
    channel_invite_url = State()
    brand_gender = State()
    brand_category = State()
    brand_name = State()
    brand_emoji = State()
    report_channel_id = State()
    report_channel_url = State()


@router.callback_query(F.data == "mgr:logout")
async def cb_logout(call: CallbackQuery, state: FSMContext, repos: Repos):
    await repos.staff.delete_session(call.from_user.id)
    await state.clear()
    await safe_edit(call.message, "👋 Вы вышли из панели менеджера.")
    await call.answer()


@router.callback_query(F.data == "mgr:back")
async def cb_back(call: CallbackQuery, state: FSMContext, repos: Repos):
    if not await require_role(call.from_user.id, repos, "manager"):
        await call.answer("❌ Сессия истекла", show_alert=True)
        return
    await state.clear()
    await safe_edit(call.message, "🧑‍💼 <b>Панель менеджера</b>", reply_markup=kb_manager())
    await call.answer()


# ── Статистика ────────────────────────────────────────────────────────

@router.callback_query(F.data == "mgr:stats")
async def cb_stats(call: CallbackQuery, repos: Repos):
    if not await require_role(call.from_user.id, repos, "manager"):
        await call.answer("❌ Сессия истекла", show_alert=True)
        return

    s = await repos.users.stats()
    o = await repos.orders.stats()
    p = await repos.posts.stats()

    admins = await repos.staff.list_by_role("admin")
    ticket_stats = await repos.tickets.all_admin_stats_today([a["login"] for a in admins])
    ticket_lines = "\n".join(
        f"  👤 <b>{t['login']}</b>: взято {t['taken']} / закрыто {t['closed']} / ожидание {t['waiting']} | 👍{t['satisfied']} 👎{t['unsatisfied']}"
        for t in ticket_stats
    ) or "  —"

    text = (
        "📊 <b>Общая статистика</b>\n\n"
        f"👥 Пользователи: <b>{s['total']}</b> (👨{s['males']} 👩{s['females']})\n"
        f"  Сегодня: {s['today']} / Неделя: {s['week']} / Месяц: {s['month']}\n\n"
        f"📦 Публикации: <b>{p['total']}</b>\n"
        f"  Сегодня: {p['today']} / Неделя: {p['week']} / Месяц: {p['month']}\n\n"
        f"🛍 Заказы: <b>{o['total']}</b>\n"
        f"  Сегодня: {o['today']} / Неделя: {o['week']} / Месяц: {o['month']}\n\n"
        f"🎧 Поддержка сегодня:\n{ticket_lines}"
    )
    await safe_edit(call.message, text, reply_markup=kb_back_manager())
    await call.answer()


# ── Финансы ───────────────────────────────────────────────────────────

@router.callback_query(F.data == "mgr:finance")
async def cb_finance(call: CallbackQuery, repos: Repos):
    if not await require_role(call.from_user.id, repos, "manager"):
        await call.answer("❌ Сессия истекла", show_alert=True)
        return

    today = await repos.finance.summary(days=1)
    week = await repos.finance.summary(days=7)
    month = await repos.finance.summary(days=30)
    total = await repos.finance.summary()

    def fmt(s):
        return f"выручка {s['revenue']:.2f} / прибыль {s['profit']:.2f} ({s['entries']} зап.)"

    text = (
        "💰 <b>Финансовая аналитика</b>\n\n"
        f"Сегодня: {fmt(today)}\n"
        f"Неделя: {fmt(week)}\n"
        f"Месяц: {fmt(month)}\n"
        f"Всего: {fmt(total)}\n\n"
        "<i>Записи добавляются складом при оформлении фотоотчёта о выполненном заказе.</i>"
    )
    await safe_edit(call.message, text, reply_markup=kb_back_manager())
    await call.answer()


# ── Экспорт в Excel ─────────────────────────────────────────────────

@router.callback_query(F.data == "mgr:export")
async def cb_export(call: CallbackQuery, repos: Repos):
    if not await require_role(call.from_user.id, repos, "manager"):
        await call.answer("❌ Сессия истекла", show_alert=True)
        return

    await call.answer("⏳ Генерирую файл...")

    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()

    ws_users = wb.active
    ws_users.title = "Пользователи"
    headers_u = ["ID", "TG ID", "Username", "Телефон", "Пол", "Язык", "Дата регистрации"]
    for col, h in enumerate(headers_u, 1):
        c = ws_users.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2C3E50")

    users = await repos.users.all()
    for i, u in enumerate(users, 2):
        ws_users.cell(row=i, column=1, value=i - 1)
        ws_users.cell(row=i, column=2, value=u.get("tg_id"))
        ws_users.cell(row=i, column=3, value=u.get("username") or "—")
        ws_users.cell(row=i, column=4, value=u.get("phone") or "—")
        ws_users.cell(row=i, column=5, value="Мужской" if u.get("gender") == "male" else "Женский")
        ws_users.cell(row=i, column=6, value=(u.get("language") or "?").upper())
        ws_users.cell(row=i, column=7, value=str(u.get("registered_at", "")))

    ws_fin = wb.create_sheet("Финансы")
    summary = await repos.finance.summary()
    rows = [("Выручка", summary["revenue"]), ("Себестоимость", summary["cost"]), ("Прибыль", summary["profit"])]
    for i, (k, v) in enumerate(rows, 1):
        ws_fin.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws_fin.cell(row=i, column=2, value=float(v))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"themoda_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    await call.message.answer_document(BufferedInputFile(buf.read(), filename=filename), caption="📥 Отчёт готов!")


# ── Тексты и баннер ──────────────────────────────────────────────────

@router.callback_query(F.data == "mgr:texts")
async def cb_texts(call: CallbackQuery, repos: Repos):
    if not await require_role(call.from_user.id, repos, "manager"):
        await call.answer("❌ Сессия истекла", show_alert=True)
        return
    await safe_edit(call.message, "✏️ Настройка текстов и баннеров находится в разработке.", reply_markup=kb_back_manager())
    await call.answer()


# ── Словарь красивых названий категорий ──────────────────────────────

CATEGORY_LABELS_FLAT = {
    ("male", "clothes"): "👔 Муж. одежда", ("male", "shoes"): "👟 Муж. обувь",
    ("male", "accessories"): "⌚️ Муж. аксессуары",
    ("female", "clothes"): "👗 Жен. одежда", ("female", "shoes"): "👠 Жен. обувь",
    ("female", "accessories"): "💍 Жен. аксессуары", ("female", "bags"): "👜 Жен. сумки",
}


# ── Каналы (муж/жен) ────────────────────────────────────────────────

def kb_channels_menu() -> InlineKeyboardMarkup:
    rows = []
    for gender in ("male", "female"):
        categories_list = CATEGORIES.get(gender, []) if isinstance(CATEGORIES, dict) else CATEGORIES
        for key in categories_list:
            label = CATEGORY_LABELS_FLAT.get((gender, key), key)
            rows.append([InlineKeyboardButton(text=label, callback_data=f"mgr:channel_set:{gender}:{key}")])
    rows.append([InlineKeyboardButton(text="⬅️ Назад", callback_data="mgr:back")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


@router.callback_query(F.data == "mgr:overview")
async def cb_overview(call: CallbackQuery, repos: Repos):
    if not await require_role(call.from_user.id, repos, "manager"):
        await call.answer("❌ Сессия истекла", show_alert=True)
        return

    channels = await repos.brands.list_channels()
    by_key = {(c["gender"], c["category"]): c for c in channels}
    brands = await repos.brands.list_all_brands()

    lines = ["👁 <b>Каналы и разделы</b>\n"]

    for gender in ("male", "female"):
        categories_list = CATEGORIES.get(gender, []) if isinstance(CATEGORIES, dict) else CATEGORIES
        for key in categories_list:
            ch = by_key.get((gender, key))
            cat_label = CATEGORY_LABELS_FLAT.get((gender, key), key)
            lines.append(f"\n<b>{cat_label}</b>")
            if ch:
                lines.append(f"  📡 {ch['chat_id']}")
            else:
                lines.append("  📡 канал не настроен")

            cat_brands = [b for b in brands if b["gender"] == gender and b["category"] == key]
            if cat_brands:
                for b in cat_brands:
                    lines.append(f"  • {b['emoji']} {b['name']} (topic_id={b['topic_id']})")
            else:
                lines.append("  • разделов пока нет")

    report_male = await repos.settings.get("report_channel_male_chat_id")
    report_female = await repos.settings.get("report_channel_female_chat_id")
    lines.append(
        f"\n<b>📦 Каналы отчётов склада</b>\n"
        f"  👨 {report_male or 'не настроен'}\n  👩 {report_female or 'не настроен'}"
    )

    text = "\n".join(lines)
    if len(text) > 4000:
        text = text[:4000] + "\n…"

    await safe_edit(call.message, text, reply_markup=kb_back_manager())
    await call.answer()


# ── Канал отчётов склада ────────────────────────────────────────────

def kb_report_channel_gender() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="👨 Мужской", callback_data="mgr:report_gender:male")],
        [InlineKeyboardButton(text="👩 Женский", callback_data="mgr:report_gender:female")],
        [InlineKeyboardButton(text="⬅️ Назад", callback_data="mgr:back")],
    ])


@router.callback_query(F.data == "mgr:report_channel")
async def cb_report_channel(call: CallbackQuery, repos: Repos):
    if not await require_role(call.from_user.id, repos, "manager"):
        await call.answer("❌ Сессия истекла", show_alert=True)
        return

    male_id = await repos.settings.get("report_channel_male_chat_id")
    female_id = await repos.settings.get("report_channel_female_chat_id")
    text = (
        "📦 <b>Каналы отчётов склада</b>\n\n"
        f"👨 Мужской: <code>{male_id or 'не настроен'}</code>\n"
        f"👩 Женский: <code>{female_id or 'не настроен'}</code>\n\n"
        "Выберите, какой канал настроить:"
    )
    await safe_edit(call.message, text, reply_markup=kb_report_channel_gender())
    await call.answer()


@router.callback_query(F.data.startswith("mgr:report_gender:"))
async def cb_report_gender(call: CallbackQuery, state: FSMContext, repos: Repos):
    if not await require_role(call.from_user.id, repos, "manager"):
        await call.answer("❌ Сессия истекла", show_alert=True)
        return

    gender = call.data.split(":")[2]
    await state.update_data(report_gender=gender)
    await state.set_state(ManagerStates.report_channel_id)
    label = "Мужской" if gender == "male" else "Женский"
    await safe_edit(call.message,
        f"📦 <b>Канал отчётов склада — {label}</b>\n\n"
        "Введите ID или @username канала, куда склад будет публиковать фотоотчёты "
        "по товарам этого пола:",
        reply_markup=kb_back_manager(),
    )
    await call.answer()


@router.message(ManagerStates.report_channel_id)
async def msg_report_channel_id(message: Message, state: FSMContext, repos: Repos):
    if not await require_role(message.from_user.id, repos, "manager"):
        await state.clear()
        return

    value = message.text.strip()
    is_numeric = value.lstrip("-").isdigit()
    is_username = value.startswith("@") and len(value) > 1
    if not (is_numeric or is_username):
        await message.answer("⚠️ Введите числовой ID (-100...) или @username канала:")
        return

    data = await state.get_data()
    gender = data.get("report_gender", "male")
    await repos.settings.set(f"report_channel_{gender}_chat_id", value)
    await state.update_data(report_chat_id=value)
    await state.set_state(ManagerStates.report_channel_url)
    await message.answer(
        "Теперь введите публичную ссылку на этот канал — она будет показана клиентам "
        "как кнопка «🔥 Витрина пополнений» (например https://t.me/themodawarehousemen):"
    )


@router.message(ManagerStates.report_channel_url)
async def msg_report_channel_url(message: Message, state: FSMContext, repos: Repos):
    if not await require_role(message.from_user.id, repos, "manager"):
        await state.clear()
        return

    data = await state.get_data()
    gender = data.get("report_gender", "male")
    url = message.text.strip()
    await repos.settings.set(f"report_channel_{gender}_url", url)
    await state.clear()

    label = "Мужской" if gender == "male" else "Женский"
    await message.answer(
        f"✅ Канал отчётов склада ({label}) сохранён!\n"
        f"ID: <code>{data.get('report_chat_id')}</code>\nСсылка: {url}\n\n"
        "💡 Не забудь добавить бота админом в этот канал.",
        reply_markup=kb_manager(),
    )


@router.callback_query(F.data == "mgr:channels")
async def cb_channels(call: CallbackQuery, repos: Repos):
    if not await require_role(call.from_user.id, repos, "manager"):
        await call.answer("❌ Сессия истекла", show_alert=True)
        return

    channels = await repos.brands.list_channels()
    by_key = {(c["gender"], c["category"]): c for c in channels}

    lines = []
    for gender in ("male", "female"):
        categories_list = CATEGORIES.get(gender, []) if isinstance(CATEGORIES, dict) else CATEGORIES
        for key in categories_list:
            c = by_key.get((gender, key))
            label = CATEGORY_LABELS_FLAT.get((gender, key), key)
            lines.append(f"{label}: {c['chat_id'] if c else '—'}")

    text = "📡 <b>Каналы (пол × категория)</b>\n\n" + "\n".join(lines)
    await safe_edit(call.message, text, reply_markup=kb_channels_menu())
    await call.answer()


@router.callback_query(F.data.startswith("mgr:channel_set:"))
async def cb_channel_set(call: CallbackQuery, state: FSMContext, repos: Repos):
    if not await require_role(call.from_user.id, repos, "manager"):
        await call.answer("❌ Сессия истекла", show_alert=True)
        return
    _, _, gender, category = call.data.split(":")
    await state.update_data(channel_gender=gender, channel_category=category)
    await state.set_state(ManagerStates.channel_chat_id)
    label = CATEGORY_LABELS_FLAT.get((gender, category), category)
    await safe_edit(call.message,
        f"Настройка канала: <b>{label}</b>\n\n"
        "Введите ID или username супергруппы:\n"
        "• Для приватной группы — числовой ID (формат -100xxxxxxxxxx)\n"
        "• Для публичной группы — @username (например @themodaclothes)\n\n"
        "💡 Бот должен быть админом в группе, форум-режим (топики) включён."
    )
    await call.answer()


@router.message(ManagerStates.channel_chat_id)
async def msg_channel_chat_id(message: Message, state: FSMContext, repos: Repos):
    if not await require_role(message.from_user.id, repos, "manager"):
        await state.clear()
        return
    chat_id = message.text.strip()
    is_numeric = chat_id.lstrip("-").isdigit()
    is_username = chat_id.startswith("@") and len(chat_id) > 1
    if not (is_numeric or is_username):
        await message.answer("⚠️ Введите числовой ID (-100...) или @username группы. Попробуйте снова:")
        return
    await state.update_data(channel_chat_id=chat_id)
    await state.set_state(ManagerStates.channel_invite_url)
    await message.answer(
        "Введите ссылку-базу для топиков.\n"
        "• Приватная группа: https://t.me/c/1234567890 (без номера топика в конце)\n"
        "• Публичная группа: https://t.me/themodaclothes (без номера топика в конце)"
    )


@router.message(ManagerStates.channel_invite_url)
async def msg_channel_invite_url(message: Message, state: FSMContext, repos: Repos):
    if not await require_role(message.from_user.id, repos, "manager"):
        await state.clear()
        return
    data = await state.get_data()
    await repos.brands.set_channel(
        data["channel_gender"], data["channel_category"], data["channel_chat_id"], message.text.strip()
    )
    await state.clear()
    await message.answer("✅ Канал сохранён!", reply_markup=kb_manager())


# ── Бренды (топики) ───────────────────────────────────────────────────

def kb_brand_gender() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="👨 Мужской", callback_data="brand_new:male")],
        [InlineKeyboardButton(text="👩 Женский", callback_data="brand_new:female")],
        [InlineKeyboardButton(text="⬅️ Назад", callback_data="mgr:brands")],
    ])


def kb_brand_category(gender: str) -> InlineKeyboardMarkup:
    categories_list = CATEGORIES.get(gender, []) if isinstance(CATEGORIES, dict) else CATEGORIES
    rows = [[InlineKeyboardButton(text=CATEGORY_LABELS_FLAT.get((gender, key), key), callback_data=f"brand_cat:{gender}:{key}")]
            for key in categories_list]
    rows.append([InlineKeyboardButton(text="⬅️ Назад", callback_data="mgr:brands")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


@router.callback_query(F.data == "mgr:brands")
async def cb_brands_menu(call: CallbackQuery, state: FSMContext, repos: Repos):
    if not await require_role(call.from_user.id, repos, "manager"):
        await call.answer("❌ Сессия истекла", show_alert=True)
        return
    await state.clear()
    await safe_edit(call.message,
        "📂 <b>Разделы</b>\n\nКаждый раздел — топик в канале (муж/жен). Раздел может быть брендом (для обуви) или типом товара (для одежды/аксессуаров).",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="➕ Добавить раздел", callback_data="brand_add")],
            [InlineKeyboardButton(text="🗑 Удалить раздел", callback_data="brand_del_menu")],
            [InlineKeyboardButton(text="⬅️ Назад", callback_data="mgr:back")],
        ]),
    )
    await call.answer()


@router.callback_query(F.data == "brand_add")
async def cb_brand_add(call: CallbackQuery, repos: Repos):
    if not await require_role(call.from_user.id, repos, "manager"):
        await call.answer("❌ Сессия истекла", show_alert=True)
        return
    await safe_edit(call.message, "➕ Выберите пол:", reply_markup=kb_brand_gender())
    await call.answer()


@router.callback_query(F.data.startswith("brand_new:"))
async def cb_brand_gender(call: CallbackQuery, repos: Repos):
    if not await require_role(call.from_user.id, repos, "manager"):
        await call.answer("❌ Сессия истекла", show_alert=True)
        return
    gender = call.data.split(":")[1]
    await safe_edit(call.message, "Выберите категорию:", reply_markup=kb_brand_category(gender))
    await call.answer()


@router.callback_query(F.data.startswith("brand_cat:"))
async def cb_brand_category(call: CallbackQuery, state: FSMContext, repos: Repos):
    if not await require_role(call.from_user.id, repos, "manager"):
        await call.answer("❌ Сессия истекла", show_alert=True)
        return
    _, gender, category = call.data.split(":")
    await state.update_data(brand_gender=gender, brand_category=category)
    await state.set_state(ManagerStates.brand_name)
    await safe_edit(call.message, "Введите название раздела (например GUCCI, или Куртки, или Часы):")
    await call.answer()


@router.message(ManagerStates.brand_name)
async def msg_brand_name(message: Message, state: FSMContext, repos: Repos):
    if not await require_role(message.from_user.id, repos, "manager"):
        await state.clear()
        return
    await state.update_data(brand_name=message.text.strip().upper())
    await state.set_state(ManagerStates.brand_emoji)
    await message.answer("Введите эмодзи раздела (например 👔):")


@router.message(ManagerStates.brand_emoji)
async def msg_brand_emoji(message: Message, state: FSMContext, repos: Repos, bot: Bot):
    if not await require_role(message.from_user.id, repos, "manager"):
        await state.clear()
        return

    emoji = message.text.strip()
    data = await state.get_data()
    gender = data["brand_gender"]
    category = data["brand_category"]
    name = data["brand_name"]
    await state.clear()

    channel = await repos.brands.get_channel(gender, category)
    if not channel:
        cat_label = CATEGORY_LABELS_FLAT.get((gender, category), category)
        await message.answer(
            f"❌ Для раздела «{cat_label}» сначала нужно настроить канал.\n"
            "Зайди в «📡 Каналы» и укажи chat_id + ссылку для этой категории, потом повтори создание раздела.",
            reply_markup=kb_manager(),
        )
        return

    try:
        topic = await bot.create_forum_topic(chat_id=channel["chat_id"], name=f"{emoji} {name}")
    except Exception as e:
        await message.answer(
            f"❌ Не удалось создать топик в группе: <code>{e}</code>\n\n"
            "Проверь:\n"
            "• Бот добавлен админом в эту группу\n"
            "• У бота включено право «Manage Topics» (Управление темами)\n"
            "• Группа реально супергруппа с включённым режимом Topics\n"
            "• chat_id в «📡 Каналы» указан верно",
            reply_markup=kb_manager(),
        )
        return

    brand_id = await repos.brands.create(
        name=name, emoji=emoji, gender=gender, category=category,
        topic_id=topic.message_thread_id,
    )
    cat_label = CATEGORY_LABELS_FLAT.get((gender, category), category)
    await message.answer(
        f"✅ <b>Раздел создан, топик в группе создан автоматически!</b>\n\n"
        f"{emoji} <b>{name}</b>\n📂 {cat_label}\n🆔 topic_id: {topic.message_thread_id}",
        reply_markup=kb_manager(),
    )


# ── Удаление разделов (топиков) ─────────────────────────────────────────

def kb_brand_del_gender() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="👨 Мужской", callback_data="brand_delg:male")],
        [InlineKeyboardButton(text="👩 Женский", callback_data="brand_delg:female")],
        [InlineKeyboardButton(text="⬅️ Назад", callback_data="mgr:brands")],
    ])


def kb_brand_del_category(gender: str) -> InlineKeyboardMarkup:
    categories_list = CATEGORIES.get(gender, []) if isinstance(CATEGORIES, dict) else CATEGORIES
    rows = [[InlineKeyboardButton(text=CATEGORY_LABELS_FLAT.get((gender, key), key), callback_data=f"brand_delc:{gender}:{key}")]
            for key in categories_list]
    rows.append([InlineKeyboardButton(text="⬅️ Назад", callback_data="brand_del_menu")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def kb_brand_del_pick(gender: str, category: str, brands: list[dict]) -> InlineKeyboardMarkup:
    rows = [[InlineKeyboardButton(text=f"{b['emoji']} {b['name']}", callback_data=f"brand_delpick:{b['id']}")]
            for b in brands]
    rows.append([InlineKeyboardButton(text="⬅️ Назад", callback_data=f"brand_delg:{gender}")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def kb_brand_del_confirm(brand_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✅ Да, удалить", callback_data=f"brand_delyes:{brand_id}")],
        [InlineKeyboardButton(text="❌ Отмена", callback_data="mgr:brands")],
    ])


@router.callback_query(F.data == "brand_del_menu")
async def cb_brand_del_menu(call: CallbackQuery, repos: Repos):
    if not await require_role(call.from_user.id, repos, "manager"):
        await call.answer("❌ Сессия истекла", show_alert=True)
        return
    await safe_edit(call.message, "🗑 Удаление раздела\n\nВыберите пол:", reply_markup=kb_brand_del_gender())
    await call.answer()


@router.callback_query(F.data.startswith("brand_delg:"))
async def cb_brand_del_gender(call: CallbackQuery, repos: Repos):
    if not await require_role(call.from_user.id, repos, "manager"):
        await call.answer("❌ Сессия истекла", show_alert=True)
        return
    gender = call.data.split(":")[1]
    await safe_edit(call.message, "Выберите категорию:", reply_markup=kb_brand_del_category(gender))
    await call.answer()


@router.callback_query(F.data.startswith("brand_delc:"))
async def cb_brand_del_category(call: CallbackQuery, repos: Repos):
    if not await require_role(call.from_user.id, repos, "manager"):
        await call.answer("❌ Сессия истекла", show_alert=True)
        return
    _, gender, category = call.data.split(":")
    brands = await repos.brands.list(gender, category)
    if not brands:
        await call.answer("В этой категории пока нет разделов.", show_alert=True)
        return
    await safe_edit(call.message, "🗑 Выберите раздел для удаления:",
                     reply_markup=kb_brand_del_pick(gender, category, brands))
    await call.answer()


@router.callback_query(F.data.startswith("brand_delpick:"))
async def cb_brand_del_pick(call: CallbackQuery, repos: Repos):
    if not await require_role(call.from_user.id, repos, "manager"):
        await call.answer("❌ Сессия истекла", show_alert=True)
        return
    brand_id = int(call.data.split(":")[1])
    brand = await repos.brands.get(brand_id)
    if not brand:
        await call.answer("Раздел не найден.", show_alert=True)
        return
    await safe_edit(
        call.message,
        f"⚠️ Удалить раздел «{brand['emoji']} {brand['name']}»?\n\n"
        "Товары этого раздела останутся в истории заказов, но раздел исчезнет из каталога и списков.\n"
        "Топик в канале Telegram, если получится, тоже будет закрыт/удалён.",
        reply_markup=kb_brand_del_confirm(brand_id),
    )
    await call.answer()


@router.callback_query(F.data.startswith("brand_delyes:"))
async def cb_brand_del_confirm(call: CallbackQuery, repos: Repos, bot: Bot):
    if not await require_role(call.from_user.id, repos, "manager"):
        await call.answer("❌ Сессия истекла", show_alert=True)
        return
    brand_id = int(call.data.split(":")[1])
    brand = await repos.brands.get(brand_id)
    if not brand:
        await call.answer("Раздел не найден.", show_alert=True)
        return

    channel = await repos.brands.get_channel(brand["gender"], brand["category"])
    if channel and brand.get("topic_id"):
        try:
            await bot.delete_forum_topic(chat_id=channel["chat_id"], message_thread_id=brand["topic_id"])
        except Exception:
            pass

    await repos.brands.deactivate(brand_id)
    await call.message.answer(
        f"✅ Раздел «{brand['emoji']} {brand['name']}» удалён.",
        reply_markup=kb_manager(),
    )
    await call.answer()


# ── Создание, удаление и просмотр стафф-аккаунтов ─────────────────────

def kb_role_select() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🛠 Админ", callback_data="acc_role:admin")],
        [InlineKeyboardButton(text="📦 Склад", callback_data="acc_role:warehouse")],
        [InlineKeyboardButton(text="🤝 Партнёр", callback_data="acc_role:partner")],
        [InlineKeyboardButton(text="⬅️ Назад", callback_data="mgr:back")],
    ])


def kb_delete_role_select() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🛠 Админ", callback_data="del_role:admin")],
        [InlineKeyboardButton(text="📦 Склад", callback_data="del_role:warehouse")],
        [InlineKeyboardButton(text="🤝 Партнёр", callback_data="del_role:partner")],
        [InlineKeyboardButton(text="⬅️ Назад", callback_data="mgr:back")],
    ])


@router.callback_query(F.data == "mgr:list_accounts")
async def cb_list_accounts(call: CallbackQuery, repos: Repos):
    if not await require_role(call.from_user.id, repos, "manager"):
        await call.answer("❌ Сессия истекла", show_alert=True)
        return

    admins = await repos.staff.list_by_role("admin")
    warehouse = await repos.staff.list_by_role("warehouse")
    partners = await repos.staff.list_partners()

    def format_list(items, is_partner=False):
        if not items:
            return "  — нет"
        res = []
        for it in items:
            login = it.get("login")
            if is_partner:
                comm = it.get("commission_pct", 0)
                res.append(f"  • <code>{login}</code> (комиссия: {comm}%)")
            else:
                res.append(f"  • <code>{login}</code>")
        return "\n".join(res)

    text = (
        "👥 <b>Список аккаунтов сотрудников</b>\n\n"
        f"🛠 <b>Администраторы:</b>\n{format_list(admins)}\n\n"
        f"📦 <b>Склад:</b>\n{format_list(warehouse)}\n\n"
        f"🤝 <b>Партнёры:</b>\n{format_list(partners, is_partner=True)}"
    )
    await safe_edit(call.message, text, reply_markup=kb_back_manager())
    await call.answer()


@router.callback_query(F.data == "mgr:delete_account")
async def cb_delete_account_menu(call: CallbackQuery, repos: Repos):
    if not await require_role(call.from_user.id, repos, "manager"):
        await call.answer("❌ Сессия истекла", show_alert=True)
        return
    await safe_edit(call.message, "🗑 <b>Удаление аккаунта</b>\n\nВыберите роль:", reply_markup=kb_delete_role_select())
    await call.answer()


@router.callback_query(F.data.startswith("del_role:"))
async def cb_delete_role(call: CallbackQuery, repos: Repos):
    if not await require_role(call.from_user.id, repos, "manager"):
        await call.answer("❌ Сессия истекла", show_alert=True)
        return

    role = call.data.split(":")[1]
    if role == "partner":
        accounts = await repos.staff.list_partners()
    else:
        accounts = await repos.staff.list_by_role(role)

    if not accounts:
        await call.answer("Аккаунтов этой роли нет.", show_alert=True)
        return

    rows = [[InlineKeyboardButton(text=f"🗑 {a['login']}", callback_data=f"del_acc:{role}:{a['login']}")] for a in accounts]
    rows.append([InlineKeyboardButton(text="⬅️ Назад", callback_data="mgr:delete_account")])
    await safe_edit(call.message, "🗑 Выберите аккаунт для удаления:", reply_markup=InlineKeyboardMarkup(inline_keyboard=rows))
    await call.answer()


@router.callback_query(F.data.startswith("del_acc:"))
async def cb_delete_account_confirm_ask(call: CallbackQuery, repos: Repos):
    if not await require_role(call.from_user.id, repos, "manager"):
        await call.answer("❌ Сессия истекла", show_alert=True)
        return

    _, role, login = call.data.split(":")
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✅ Да, удалить", callback_data=f"del_acc_yes:{role}:{login}")],
        [InlineKeyboardButton(text="❌ Отмена", callback_data=f"del_role:{role}")],
    ])
    await safe_edit(
        call.message,
        f"⚠️ Удалить аккаунт <code>{login}</code> ({role})?\nЭто действие нельзя отменить.",
        reply_markup=kb,
    )
    await call.answer()


@router.callback_query(F.data.startswith("del_acc_yes:"))
async def cb_delete_account_do(call: CallbackQuery, repos: Repos):
    if not await require_role(call.from_user.id, repos, "manager"):
        await call.answer("❌ Сессия истекла", show_alert=True)
        return

    _, role, login = call.data.split(":")
    ok = await repos.staff.delete_account(role, login)

    if ok:
        await safe_edit(call.message, f"✅ Аккаунт <code>{login}</code> ({role}) удалён.", reply_markup=kb_back_manager())
    else:
        await safe_edit(call.message, f"❌ Не удалось удалить <code>{login}</code>.", reply_markup=kb_back_manager())
    await call.answer()


@router.callback_query(F.data == "mgr:create_account")
async def cb_create_account(call: CallbackQuery, repos: Repos):
    if not await require_role(call.from_user.id, repos, "manager"):
        await call.answer("❌ Сессия истекла", show_alert=True)
        return
    await safe_edit(call.message, "➕ Выберите роль для нового аккаунта:", reply_markup=kb_role_select())
    await call.answer()


@router.callback_query(F.data.startswith("acc_role:"))
async def cb_acc_role(call: CallbackQuery, state: FSMContext, repos: Repos):
    if not await require_role(call.from_user.id, repos, "manager"):
        await call.answer("❌ Сессия истекла", show_alert=True)
        return
    role = call.data.split(":")[1]
    await state.update_data(new_role=role)
    await state.set_state(ManagerStates.create_login)
    await safe_edit(call.message, f"➕ Создание аккаунта ({role})\n\nВведите логин для нового пользователя:")
    await call.answer()


@router.message(ManagerStates.create_login)
async def msg_create_login(message: Message, state: FSMContext, repos: Repos):
    if not await require_role(message.from_user.id, repos, "manager"):
        await state.clear()
        return

    login = message.text.strip()
    if not login:
        await message.answer("⚠️ Логин не может быть пустым. Введите логин:")
        return

    await state.update_data(new_login=login)
    await state.set_state(ManagerStates.create_password)
    await message.answer("Введите пароль для нового аккаунта:")


@router.message(ManagerStates.create_password)
async def msg_create_password(message: Message, state: FSMContext, repos: Repos):
    if not await require_role(message.from_user.id, repos, "manager"):
        await state.clear()
        return

    password = message.text.strip()
    if not password:
        await message.answer("⚠️ Пароль не может быть пустым. Введите пароль:")
        return

    await state.update_data(new_password=password)
    data = await state.get_data()
    role = data.get("new_role")

    # Если создаем партнера, запрашиваем процент комиссии, иначе завершаем создание
    if role == "partner":
        await state.set_state(ManagerStates.create_commission)
        await message.answer("Введите процент комиссии для партнера (например, 10 или 15.5):")
    else:
        await finalize_account_creation(message, state, repos, commission=0)


@router.message(ManagerStates.create_commission)
async def msg_create_commission(message: Message, state: FSMContext, repos: Repos):
    if not await require_role(message.from_user.id, repos, "manager"):
        await state.clear()
        return

    try:
        commission = float(message.text.strip().replace(",", "."))
    except ValueError:
        await message.answer("⚠️ Введите корректное число для комиссии (например 10):")
        return

    await finalize_account_creation(message, state, repos, commission=commission)


async def finalize_account_creation(message: Message, state: FSMContext, repos: Repos, commission: float):
    data = await state.get_data()
    role = data.get("new_role")
    login = data.get("new_login")
    password = data.get("new_password")
    await state.clear()

    try:
        if role == "partner":
            await repos.staff.create_partner(login=login, password=password, commission_pct=commission)
        else:
            await repos.staff.create_account(role=role, login=login, password=password)
        
        await message.answer(
            f"✅ <b>Аккаунт успешно создан!</b>\n\n"
            f"Роль: <b>{role}</b>\n"
            f"Логин: <code>{login}</code>\n"
            f"Пароль: <code>{password}</code>"
            + (f"\nКомиссия: {commission}%" if role == "partner" else ""),
            reply_markup=kb_manager(),
        )
    except Exception as e:
        await message.answer(
            f"❌ Ошибка при создании аккаунта (возможно, такой логин уже занят):\n<code>{e}</code>",
            reply_markup=kb_manager(),
        )


# ── Смена собственного пароля менеджера ───────────────────────────────

@router.callback_query(F.data == "mgr:change_my_password")
async def cb_change_my_password(call: CallbackQuery, state: FSMContext, repos: Repos):
    if not await require_role(call.from_user.id, repos, "manager"):
        await call.answer("❌ Сессия истекла", show_alert=True)
        return
    await state.set_state(ManagerStates.change_my_password)
    await safe_edit(call.message, "🔑 <b>Смена пароля</b>\n\nВведите новый пароль для своего аккаунта менеджера:", reply_markup=kb_back_manager())
    await call.answer()


@router.message(ManagerStates.change_my_password)
async def msg_change_my_password(message: Message, state: FSMContext, repos: Repos):
    if not await require_role(message.from_user.id, repos, "manager"):
        await state.clear()
        return

    new_pwd = message.text.strip()
    if not new_pwd:
        await message.answer("⚠️ Пароль не может быть пустым. Введите новый пароль:")
        return

    await repos.staff.update_password("manager", "manager", new_pwd)  # или через сессию
    await state.clear()
    await message.answer("✅ Ваш пароль успешно изменён!", reply_markup=kb_manager())
