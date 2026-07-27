"""
handlers/staff/partner.py

Партнёр — отдельная стафф-роль:
  /partner — вход (логин/пароль, создаёт менеджер)
  🔗 Моя ссылка       — реф-ссылка вида t.me/<bot>?start=ref_<code>
  📊 Моя статистика   — сколько перешло, сколько купило, выручка, комиссия
  📥 Экспорт в Excel   — отчёт по своим заказам/комиссии
"""
from io import BytesIO
from datetime import datetime

from aiogram import Router, F, Bot
from aiogram.types import (
    Message, CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton, BufferedInputFile,
)

from db.repos import Repos
from handlers.staff.auth import require_role
from utils import safe_edit

router = Router()


def kb_partner_menu() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔗 Моя ссылка", callback_data="ptn:link")],
        [InlineKeyboardButton(text="📊 Моя статистика", callback_data="ptn:stats")],
        [InlineKeyboardButton(text="📥 Экспорт в Excel", callback_data="ptn:export")],
        [InlineKeyboardButton(text="🚪 Выйти", callback_data="ptn:logout")],
    ])


def kb_back_partner() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="⬅️ Назад", callback_data="ptn:menu")]])


async def show_partner_panel(message: Message, session: dict):
    await message.answer(
        f"🤝 <b>Панель партнёра</b>\n👤 {session.get('login','')}",
        reply_markup=kb_partner_menu(),
    )


@router.callback_query(F.data == "ptn:menu")
async def cb_ptn_menu(call: CallbackQuery, repos: Repos):
    session = await require_role(call.from_user.id, repos, "partner")
    if not session:
        await call.answer("❌ Сессия истекла", show_alert=True)
        return
    await safe_edit(call.message, f"🤝 <b>Панель партнёра</b>\n👤 {session.get('login','')}", reply_markup=kb_partner_menu())
    await call.answer()


@router.callback_query(F.data == "ptn:logout")
async def cb_ptn_logout(call: CallbackQuery, repos: Repos):
    await repos.staff.delete_session(call.from_user.id)
    await safe_edit(call.message, "👋 Вы вышли из панели партнёра.")
    await call.answer()


@router.callback_query(F.data == "ptn:link")
async def cb_ptn_link(call: CallbackQuery, repos: Repos, bot: Bot):
    session = await require_role(call.from_user.id, repos, "partner")
    if not session:
        await call.answer("❌ Сессия истекла", show_alert=True)
        return

    account = await repos.staff.get_account("partner", session["login"])
    ref_code = account.get("ref_code") if account else None
    if not ref_code:
        await call.answer("❌ Реф-код не найден, обратитесь к менеджеру.", show_alert=True)
        return

    bot_me = await bot.get_me()
    link = f"https://t.me/{bot_me.username}?start=ref_{ref_code}"
    await safe_edit(
        call.message,
        f"🔗 <b>Ваша реферальная ссылка:</b>\n\n<code>{link}</code>\n\n"
        "Отправляйте её клиентам — все, кто перейдёт по ней и оформит заказ, "
        "будут засчитаны в вашу статистику.",
        reply_markup=kb_back_partner(),
    )
    await call.answer()


@router.callback_query(F.data == "ptn:stats")
async def cb_ptn_stats(call: CallbackQuery, repos: Repos):
    session = await require_role(call.from_user.id, repos, "partner")
    if not session:
        await call.answer("❌ Сессия истекла", show_alert=True)
        return

    account = await repos.staff.get_account("partner", session["login"])
    commission_pct = float(account.get("commission_pct") or 0) if account else 0

    s = await repos.partners.stats(session["login"], commission_pct)
    text = (
        "📊 <b>Моя статистика</b>\n\n"
        f"👥 Перешло по ссылке: <b>{s['total_referrals']}</b>\n"
        f"🛍 Купили хотя бы раз: <b>{s['buyers_count']}</b>\n"
        f"📦 Засчитанных заказов: <b>{s['orders_count']}</b>\n\n"
        f"💰 Выручка с этих заказов: <b>{s['total_revenue']}</b>\n"
        f"📈 Ваш процент: <b>{commission_pct}%</b>\n"
        f"💵 Ваша комиссия: <b>{s['commission_earned']}</b>"
    )
    await safe_edit(call.message, text, reply_markup=kb_back_partner())
    await call.answer()


@router.callback_query(F.data == "ptn:export")
async def cb_ptn_export(call: CallbackQuery, repos: Repos):
    session = await require_role(call.from_user.id, repos, "partner")
    if not session:
        await call.answer("❌ Сессия истекла", show_alert=True)
        return

    await call.answer("⏳ Генерирую файл...")

    account = await repos.staff.get_account("partner", session["login"])
    commission_pct = float(account.get("commission_pct") or 0) if account else 0
    orders = await repos.partners.orders_for_partner(session["login"])

    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заказы"

    headers = ["№ заказа", "Клиент", "Дата", "Статус", "Товары", "Сумма заказа", f"Комиссия ({commission_pct}%)"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2C3E50")

    total_revenue = 0.0
    for i, o in enumerate(orders, 2):
        order_sum = sum(repos.partners._parse_price(it.get("price", "")) for it in o["items"])
        total_revenue += order_sum
        items_str = ", ".join(it.get("title", "?") for it in o["items"])
        username = f"@{o['username']}" if o.get("username") else f"ID {o['user_tg_id']}"

        ws.cell(row=i, column=1, value=o["order_id"])
        ws.cell(row=i, column=2, value=username)
        ws.cell(row=i, column=3, value=str(o.get("created_at", "")))
        ws.cell(row=i, column=4, value=o.get("status", ""))
        ws.cell(row=i, column=5, value=items_str)
        ws.cell(row=i, column=6, value=order_sum)
        ws.cell(row=i, column=7, value=round(order_sum * commission_pct / 100, 2))

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22

    ws_sum = wb.create_sheet("Сводка")
    rows_sum = [
        ("Перешло по ссылке", await repos.partners.referral_count(session["login"])),
        ("Засчитанных заказов", len(orders)),
        ("Общая выручка", round(total_revenue, 2)),
        ("Ваш процент", f"{commission_pct}%"),
        ("Ваша комиссия", round(total_revenue * commission_pct / 100, 2)),
    ]
    for i, (k, v) in enumerate(rows_sum, 1):
        ws_sum.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws_sum.cell(row=i, column=2, value=v)
    ws_sum.column_dimensions["A"].width = 26
    ws_sum.column_dimensions["B"].width = 18

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"partner_{session['login']}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    await call.message.answer_document(BufferedInputFile(buf.read(), filename=filename), caption="📥 Ваш отчёт готов!")